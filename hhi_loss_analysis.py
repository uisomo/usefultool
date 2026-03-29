#!/usr/bin/env python3
"""
HHI vs Portfolio Loss Rate Analysis via Monte Carlo Simulation.

Demonstrates how portfolio concentration (measured by HHI) affects
loss distribution under purely idiosyncratic risk (zero correlation).

Conditions:
  - 10 obligors
  - PD = 0.03% (AA rating)
  - LGD = 100%
  - Correlation = 0
  - 1,000,000 simulation trials per HHI point
"""

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib import font_manager
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ── Japanese font setup ──────────────────────────────────────────────────
font_path = "/usr/share/fonts/opentype/ipafont-gothic/ipag.ttf"
font_manager.fontManager.addfont(font_path)
plt.rcParams["font.family"] = "IPAGothic"

# ── Configuration ────────────────────────────────────────────────────────
N_OBLIGORS = 10
PD = 0.0003           # 0.03% (AA rating)
LGD = 1.0             # 100%
N_SIMS = 1_000_000
N_HHI_POINTS = 50


def generate_weights(alpha, n=N_OBLIGORS):
    """Generate weight vector with concentration parameter alpha in [0, 1].

    alpha=0 -> equal weights (HHI = 1/N)
    alpha=1 -> full concentration on obligor 0 (HHI = 1.0)
    """
    equal = np.ones(n) / n
    concentrated = np.zeros(n)
    concentrated[0] = 1.0
    return (1 - alpha) * equal + alpha * concentrated


def compute_hhi(weights):
    return np.sum(weights ** 2)


def run_simulation(weights, n_sims=N_SIMS):
    """Run Monte Carlo and return portfolio loss rates array."""
    n = len(weights)
    defaults = (np.random.random((n_sims, n)) < PD).astype(np.float64)
    return defaults @ (weights * LGD)


def compute_stats(loss_rates):
    return {
        "expected_loss": np.mean(loss_rates),
        "std_dev": np.std(loss_rates),
        "var_999": np.percentile(loss_rates, 99.9),
        "var_99": np.percentile(loss_rates, 99.0),
        "max_loss": np.max(loss_rates),
    }


def theoretical_std(hhi_values):
    return np.sqrt(hhi_values * PD * (1 - PD))


def create_chart(hhi_values, stats_list):
    fig, ax = plt.subplots(figsize=(12, 8))

    hhis = np.array(hhi_values)
    std_devs = np.array([s["std_dev"] for s in stats_list])
    var_999 = np.array([s["var_999"] for s in stats_list])
    max_losses = np.array([s["max_loss"] for s in stats_list])
    expected = np.array([s["expected_loss"] for s in stats_list])

    # Simulated metrics
    ax.plot(hhis, expected * 100, "o-", label="期待損失率 (E[L])",
            markersize=3, color="#2196F3")
    ax.plot(hhis, std_devs * 100, "s-", label="標準偏差 (σ)",
            markersize=3, color="#FF9800")
    ax.plot(hhis, var_999 * 100, "^-", label="VaR 99.9%",
            markersize=4, color="#E91E63")
    ax.plot(hhis, max_losses * 100, "D-", label="最大損失率",
            markersize=3, color="#9C27B0", alpha=0.6)

    # Theoretical curves
    hhi_fine = np.linspace(hhis.min(), hhis.max(), 200)
    ax.plot(hhi_fine, theoretical_std(hhi_fine) * 100, "--",
            color="red", label="理論値 σ = √(HHI·PD·(1-PD))", linewidth=2)
    ax.axhline(y=PD * 100, color="gray", linestyle=":",
               label=f"理論値 E[L] = {PD*100:.3f}%", linewidth=1.5)

    ax.set_xlabel("HHI (ハーフィンダール・ハーシュマン指数)", fontsize=13)
    ax.set_ylabel("損失率 (%)", fontsize=13)
    ax.set_title(
        "HHI（ポートフォリオ集中度）と損失率の関係\n"
        f"債務者数={N_OBLIGORS}, PD={PD*100:.3f}% (AA格), "
        f"LGD={LGD*100:.0f}%, 相関=0, "
        f"シミュレーション={N_SIMS:,}回",
        fontsize=13,
    )
    ax.legend(fontsize=10, loc="upper left")
    ax.grid(True, alpha=0.3)
    ax.set_xlim(0.05, 1.05)

    plt.tight_layout()
    output_path = "hhi_loss_analysis.png"
    plt.savefig(output_path, dpi=150, bbox_inches="tight")
    print(f"Saved: {output_path}")


def create_excel(hhi_values, stats_list, all_weights):
    """Generate Excel workbook with data, charts, and weight breakdown."""
    HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
    SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="2F5496")
    PARAM_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    WARN_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    GOOD_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header(ws, row, max_col):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = THIN_BORDER

    wb = openpyxl.Workbook()

    # ── Sheet 1: Parameters ──────────────────────────────────────────────
    ws_param = wb.active
    ws_param.title = "Parameters"
    ws_param.sheet_properties.tabColor = "2F5496"

    ws_param["A1"] = "HHI vs 損失率 Monte Carlo分析"
    ws_param["A1"].font = TITLE_FONT
    ws_param.merge_cells("A1:C1")

    ws_param["A3"] = "パラメータ"
    ws_param["B3"] = "値"
    ws_param["C3"] = "説明"
    style_header(ws_param, 3, 3)

    params = [
        ("債務者数", N_OBLIGORS, "ポートフォリオ内の企業数"),
        ("PD (デフォルト確率)", PD, "AA格の年間デフォルト確率"),
        ("LGD (デフォルト時損失率)", LGD, "デフォルト時のエクスポージャー損失割合"),
        ("相関 (ρ)", 0.0, "債務者間の相関（今回はゼロ）"),
        ("シミュレーション回数", N_SIMS, "各HHIポイントでの試行回数"),
        ("HHIポイント数", N_HHI_POINTS, "分析するHHI水準の数"),
    ]
    for i, (name, val, desc) in enumerate(params):
        r = 4 + i
        ws_param.cell(row=r, column=1, value=name).fill = PARAM_FILL
        c = ws_param.cell(row=r, column=2, value=val)
        if isinstance(val, float) and val <= 1:
            c.number_format = "0.00%"
        else:
            c.number_format = "#,##0"
        ws_param.cell(row=r, column=3, value=desc)
        for col in range(1, 4):
            ws_param.cell(row=r, column=col).border = THIN_BORDER

    # Key insight box
    ws_param["A12"] = "重要な発見"
    ws_param["A12"].font = SUBTITLE_FONT
    insights = [
        "1. 期待損失率はHHIに関係なく一定 (E[L] = PD × LGD = 0.03%)",
        "2. 標準偏差はσ = √(HHI × PD × (1-PD)) に比例して増加",
        "3. VaR 99.9%はHHI増加で逆に低下する（VaRの落とし穴）",
        "   → 集中ポートフォリオはデフォルト確率が低いが、起きた時の損失が壊滅的",
        "4. 最大損失率はHHI増加とともに上昇 → 集中リスクの本質",
        "5. VaRだけでBB advance rateを設定すると集中リスクを見逃す",
    ]
    for i, text in enumerate(insights):
        ws_param.cell(row=13 + i, column=1, value=text)

    ws_param.column_dimensions["A"].width = 40
    ws_param.column_dimensions["B"].width = 16
    ws_param.column_dimensions["C"].width = 40

    # ── Sheet 2: Simulation Results ──────────────────────────────────────
    ws_data = wb.create_sheet("SimulationResults")
    ws_data.sheet_properties.tabColor = "00B050"

    ws_data["A1"] = "HHI vs 損失率 シミュレーション結果"
    ws_data["A1"].font = TITLE_FONT
    ws_data.merge_cells("A1:H1")

    headers = [
        "HHI", "期待損失率\n(E[L])", "標準偏差\n(σ)",
        "VaR 99%", "VaR 99.9%", "最大損失率",
        "理論値 σ\n√(HHI·PD·(1-PD))", "最大ウェイト\n(w_max)",
    ]
    for c, h in enumerate(headers, 1):
        ws_data.cell(row=3, column=c, value=h)
    style_header(ws_data, 3, len(headers))

    for i, (hhi, stats, weights) in enumerate(zip(hhi_values, stats_list, all_weights)):
        r = 4 + i
        ws_data.cell(row=r, column=1, value=hhi).number_format = "0.0000"
        ws_data.cell(row=r, column=2, value=stats["expected_loss"]).number_format = "0.0000%"
        ws_data.cell(row=r, column=3, value=stats["std_dev"]).number_format = "0.0000%"
        ws_data.cell(row=r, column=4, value=stats["var_99"]).number_format = "0.00%"
        ws_data.cell(row=r, column=5, value=stats["var_999"]).number_format = "0.00%"
        ws_data.cell(row=r, column=6, value=stats["max_loss"]).number_format = "0.00%"
        theo = np.sqrt(hhi * PD * (1 - PD))
        ws_data.cell(row=r, column=7, value=theo).number_format = "0.0000%"
        ws_data.cell(row=r, column=8, value=float(np.max(weights))).number_format = "0.00%"

        # Color VaR 99.9% cells - warn when low despite high HHI
        if hhi > 0.3 and stats["var_999"] == 0:
            ws_data.cell(row=r, column=5).fill = WARN_FILL

        for col in range(1, len(headers) + 1):
            ws_data.cell(row=r, column=col).border = THIN_BORDER
            ws_data.cell(row=r, column=col).alignment = Alignment(horizontal="center")

    last_data_row = 3 + len(hhi_values)

    for c in range(1, 9):
        ws_data.column_dimensions[get_column_letter(c)].width = 16

    # ── Chart 1: σ and theoretical σ ─────────────────────────────────────
    chart1 = LineChart()
    chart1.title = "HHI vs 標準偏差（シミュレーション vs 理論値）"
    chart1.x_axis.title = "HHI"
    chart1.y_axis.title = "損失率"
    chart1.y_axis.numFmt = "0.00%"
    chart1.width = 22
    chart1.height = 13

    cats = Reference(ws_data, min_col=1, min_row=4, max_row=last_data_row)
    # σ simulated (col 3)
    s1 = Reference(ws_data, min_col=3, min_row=3, max_row=last_data_row)
    chart1.add_data(s1, titles_from_data=True)
    # σ theoretical (col 7)
    s2 = Reference(ws_data, min_col=7, min_row=3, max_row=last_data_row)
    chart1.add_data(s2, titles_from_data=True)
    chart1.set_categories(cats)

    chart1.series[0].graphicalProperties.line.solidFill = "FF9800"
    chart1.series[0].graphicalProperties.line.width = 25000
    chart1.series[1].graphicalProperties.line.solidFill = "FF0000"
    chart1.series[1].graphicalProperties.line.dashStyle = "dash"
    chart1.series[1].graphicalProperties.line.width = 25000

    ws_data.add_chart(chart1, "A" + str(last_data_row + 3))

    # ── Chart 2: VaR 99.9% vs Max Loss ───────────────────────────────────
    chart2 = LineChart()
    chart2.title = "HHI vs VaR 99.9% vs 最大損失率（VaRの落とし穴）"
    chart2.x_axis.title = "HHI"
    chart2.y_axis.title = "損失率"
    chart2.y_axis.numFmt = "0%"
    chart2.width = 22
    chart2.height = 13

    # VaR 99.9% (col 5)
    v1 = Reference(ws_data, min_col=5, min_row=3, max_row=last_data_row)
    chart2.add_data(v1, titles_from_data=True)
    # Max loss (col 6)
    v2 = Reference(ws_data, min_col=6, min_row=3, max_row=last_data_row)
    chart2.add_data(v2, titles_from_data=True)
    chart2.set_categories(cats)

    chart2.series[0].graphicalProperties.line.solidFill = "E91E63"
    chart2.series[0].graphicalProperties.line.width = 25000
    chart2.series[1].graphicalProperties.line.solidFill = "9C27B0"
    chart2.series[1].graphicalProperties.line.width = 25000

    ws_data.add_chart(chart2, "A" + str(last_data_row + 19))

    # ── Chart 3: All metrics combined ────────────────────────────────────
    chart3 = LineChart()
    chart3.title = "HHI vs 全指標（集中度と損失率の関係）"
    chart3.x_axis.title = "HHI"
    chart3.y_axis.title = "損失率"
    chart3.y_axis.numFmt = "0%"
    chart3.width = 22
    chart3.height = 13

    for col_idx in [2, 3, 5, 6]:
        ref = Reference(ws_data, min_col=col_idx, min_row=3, max_row=last_data_row)
        chart3.add_data(ref, titles_from_data=True)
    chart3.set_categories(cats)

    colors = ["2196F3", "FF9800", "E91E63", "9C27B0"]
    for i, color in enumerate(colors):
        chart3.series[i].graphicalProperties.line.solidFill = color
        chart3.series[i].graphicalProperties.line.width = 25000

    ws_data.add_chart(chart3, "A" + str(last_data_row + 35))

    # ── Sheet 3: Weight Distribution ─────────────────────────────────────
    ws_wt = wb.create_sheet("WeightDistribution")
    ws_wt.sheet_properties.tabColor = "7030A0"

    ws_wt["A1"] = "各HHIポイントでの債務者ウェイト配分"
    ws_wt["A1"].font = TITLE_FONT
    ws_wt.merge_cells("A1:L1")

    wt_headers = ["HHI"] + [f"債務者{i+1}" for i in range(N_OBLIGORS)]
    for c, h in enumerate(wt_headers, 1):
        ws_wt.cell(row=3, column=c, value=h)
    style_header(ws_wt, 3, len(wt_headers))

    # Show every 5th point + first and last for readability
    sample_indices = sorted(set([0, len(hhi_values)-1] +
                                list(range(0, len(hhi_values), 5))))
    for row_idx, i in enumerate(sample_indices):
        r = 4 + row_idx
        ws_wt.cell(row=r, column=1, value=hhi_values[i]).number_format = "0.0000"
        ws_wt.cell(row=r, column=1).border = THIN_BORDER
        for j, w in enumerate(all_weights[i]):
            c = ws_wt.cell(row=r, column=2 + j, value=float(w))
            c.number_format = "0.00%"
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center")
            # Highlight dominant weight
            if w > 0.5:
                c.fill = WARN_FILL
            elif abs(w - 1.0/N_OBLIGORS) < 0.001:
                c.fill = GOOD_FILL

    ws_wt.column_dimensions["A"].width = 12
    for c in range(2, N_OBLIGORS + 2):
        ws_wt.column_dimensions[get_column_letter(c)].width = 12

    # Explanation
    expl_row = 4 + len(sample_indices) + 2
    ws_wt.cell(row=expl_row, column=1, value="読み方").font = SUBTITLE_FONT
    explanations = [
        "緑: 均等配分 (1/N = 10%)",
        "赤: 集中配分 (50%超のウェイト)",
        "HHI = Σ(w_i²)  均等=0.10, 完全集中=1.00",
        "HHIが高い → 1社に集中 → 名前集中リスクが高い",
    ]
    for i, text in enumerate(explanations):
        ws_wt.cell(row=expl_row + 1 + i, column=1, value=text)

    # ── Sheet 4: VaR Pitfall Explanation ─────────────────────────────────
    ws_expl = wb.create_sheet("VaRの落とし穴")
    ws_expl.sheet_properties.tabColor = "FF0000"

    ws_expl["A1"] = "VaRが集中リスクを見逃す理由"
    ws_expl["A1"].font = TITLE_FONT
    ws_expl.merge_cells("A1:E1")

    ws_expl["A3"] = "シナリオ比較"
    ws_expl["A3"].font = SUBTITLE_FONT

    comp_headers = ["指標", "均等配分\n(HHI=0.10)", "集中配分\n(HHI≈1.0)", "判定"]
    for c, h in enumerate(comp_headers, 1):
        ws_expl.cell(row=4, column=c, value=h)
    style_header(ws_expl, 4, 4)

    # Find stats for equal and most concentrated
    eq_stats = stats_list[0]
    conc_stats = stats_list[-1]

    comparisons = [
        ("期待損失率 E[L]", eq_stats["expected_loss"], conc_stats["expected_loss"], "同じ"),
        ("標準偏差 σ", eq_stats["std_dev"], conc_stats["std_dev"], "集中が高い ⚠"),
        ("VaR 99%", eq_stats["var_99"], conc_stats["var_99"], "集中が低い（罠！）"),
        ("VaR 99.9%", eq_stats["var_999"], conc_stats["var_999"], "集中が低い（罠！）"),
        ("最大損失率", eq_stats["max_loss"], conc_stats["max_loss"], "集中が圧倒的に高い ⚠⚠"),
    ]
    for i, (label, eq_val, conc_val, verdict) in enumerate(comparisons):
        r = 5 + i
        ws_expl.cell(row=r, column=1, value=label).fill = PARAM_FILL
        ws_expl.cell(row=r, column=2, value=eq_val).number_format = "0.00%"
        ws_expl.cell(row=r, column=3, value=conc_val).number_format = "0.00%"
        ws_expl.cell(row=r, column=4, value=verdict)
        if "罠" in verdict:
            ws_expl.cell(row=r, column=4).fill = WARN_FILL
        for col in range(1, 5):
            ws_expl.cell(row=r, column=col).border = THIN_BORDER
            ws_expl.cell(row=r, column=col).alignment = Alignment(horizontal="center")

    ws_expl["A12"] = "なぜVaRが低くなるか"
    ws_expl["A12"].font = SUBTITLE_FONT
    reasons = [
        "均等配分: 10社のうち1社でもデフォルトする確率 ≈ 1-(1-0.03%)^10 ≈ 0.30%",
        "  → 0.30% > 0.10%（VaR 99.9%の閾値）なのでVaRに引っかかる",
        "  → 1社デフォルト時の損失 = 10%（ウェイト均等なので）",
        "",
        "集中配分: 主要1社がデフォルトする確率 = 0.03%",
        "  → 0.03% < 0.10%（VaR 99.9%の閾値）なのでVaRに引っかからない",
        "  → しかし1社デフォルト時の損失 = 99.5%（壊滅的）",
        "",
        "結論: VaRは「確率の境界」しか見ないため、",
        "  「低確率・高影響」の集中リスクを過小評価する。",
        "  BB advance rateの設定にはExpected Shortfall(CVaR)や",
        "  ストレステストの併用が不可欠。",
    ]
    for i, text in enumerate(reasons):
        ws_expl.cell(row=13 + i, column=1, value=text)

    ws_expl.column_dimensions["A"].width = 24
    ws_expl.column_dimensions["B"].width = 18
    ws_expl.column_dimensions["C"].width = 18
    ws_expl.column_dimensions["D"].width = 28

    output_path = "hhi_loss_analysis.xlsx"
    wb.save(output_path)
    print(f"Saved: {output_path}")


def main():
    np.random.seed(42)

    alphas = np.linspace(0, 0.995, N_HHI_POINTS)
    hhi_values = []
    stats_list = []
    all_weights = []

    print(f"Running Monte Carlo: {N_HHI_POINTS} HHI points x "
          f"{N_SIMS:,} simulations each")
    print(f"  PD={PD*100:.3f}%, LGD={LGD*100:.0f}%, "
          f"N_obligors={N_OBLIGORS}, correlation=0\n")

    for i, alpha in enumerate(alphas):
        weights = generate_weights(alpha)
        hhi = compute_hhi(weights)
        hhi_values.append(hhi)
        all_weights.append(weights)

        loss_rates = run_simulation(weights)
        stats = compute_stats(loss_rates)
        stats_list.append(stats)

        if (i + 1) % 10 == 0 or i == 0:
            print(f"  [{i+1:2d}/{N_HHI_POINTS}] HHI={hhi:.4f}  "
                  f"E[L]={stats['expected_loss']*100:.4f}%  "
                  f"σ={stats['std_dev']*100:.4f}%  "
                  f"VaR99.9={stats['var_999']*100:.2f}%")

    print("\nGenerating charts...")
    create_chart(hhi_values, stats_list)
    create_excel(hhi_values, stats_list, all_weights)
    print("Done!")


if __name__ == "__main__":
    main()
