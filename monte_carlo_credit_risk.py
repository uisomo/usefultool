#!/usr/bin/env python3
"""
Monte Carlo Credit Risk Simulation - Excel Prototype Generator
Generates an Excel file with formula-based Monte Carlo simulation (no VBA).

Uses a multi-factor Gaussian copula model with per-obligor factor loadings:
  Z_i = sqrt(w_r)*F_region + sqrt(w_size)*F_size + sqrt(w_s)*F_sector
        + sqrt(w_h)*F_hnwi + sqrt(w_idio)*eps_i

Factor weights are looked up per-category from the FactorLoadings sheet:
  - Region:  US, Europe, Asia ex-Japan, Japan
  - Size:    Large, SME
  - Sector:  Finance, Manufacturing, Tech, Real Estate
  - HNWI:    Yes, No

Obligor defaults when Z_i < NORM.S.INV(PD_i)
"""

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

# ── Configuration ──────────────────────────────────────────────────────────
N_SIMS = 500
N_BUCKETS = 25  # histogram buckets

# ── Sample Obligor Data ───────────────────────────────────────────────────
OBLIGORS = [
    # (ID, Name, Region, Size, Sector, HNWI, PD%, EAD_millions, LGD%)
    ("OBL-001", "Alpha Industries",       "Japan",        "Large", "Manufacturing", "No",  1.5, 120, 45),
    ("OBL-002", "Beta Financial Group",    "Japan",        "Large", "Finance",       "No",  0.8, 200, 40),
    ("OBL-003", "Gamma Tech Solutions",    "Asia ex-Japan","SME",   "Tech",          "No",  2.0,  80, 50),
    ("OBL-004", "Delta Real Estate",       "Asia ex-Japan","SME",   "Real Estate",   "Yes", 3.0, 150, 55),
    ("OBL-005", "Epsilon Holdings",        "Japan",        "Large", "Finance",       "Yes", 1.2, 180, 40),
    ("OBL-006", "Zeta Automobil AG",       "Europe",       "Large", "Manufacturing", "No",  1.0, 160, 45),
    ("OBL-007", "Eta Banque SA",           "Europe",       "Large", "Finance",       "No",  0.5, 250, 35),
    ("OBL-008", "Theta PropCo Ltd",        "Europe",       "SME",   "Real Estate",   "No",  2.5, 100, 55),
    ("OBL-009", "Iota Digital GmbH",       "Europe",       "SME",   "Tech",          "Yes", 1.8,  90, 50),
    ("OBL-010", "Kappa Energie SA",        "Europe",       "Large", "Manufacturing", "No",  0.7, 140, 45),
    ("OBL-011", "Lambda Capital Corp",     "US",           "Large", "Finance",       "No",  0.6, 220, 38),
    ("OBL-012", "Mu Semiconductor Inc",    "US",           "SME",   "Tech",          "No",  2.2,  70, 50),
    ("OBL-013", "Nu Development LLC",      "US",           "SME",   "Real Estate",   "Yes", 3.5,  95, 58),
    ("OBL-014", "Xi Wealth Partners",      "US",           "Large", "Finance",       "Yes", 1.0, 170, 40),
    ("OBL-015", "Omicron MFG Corp",        "US",           "SME",   "Manufacturing", "Yes", 4.0,  60, 48),
]

REGIONS = ["US", "Europe", "Asia ex-Japan", "Japan"]
SIZES = ["Large", "SME"]
SECTORS = ["Finance", "Manufacturing", "Tech", "Real Estate"]
HNWI_VALS = ["No", "Yes"]

# ── Default Factor Loadings (intra-group correlations) ────────────────────
# These define how much of each obligor's risk comes from the shared factor.
# Two obligors sharing a factor have pairwise correlation = w_i + w_j + ...
REGION_LOADINGS = {"US": 0.25, "Europe": 0.20, "Asia ex-Japan": 0.15, "Japan": 0.10}
SIZE_LOADINGS = {"Large": 0.12, "SME": 0.06}
SECTOR_LOADINGS = {"Finance": 0.10, "Manufacturing": 0.08, "Tech": 0.08, "Real Estate": 0.06}
HNWI_LOADINGS = {"No": 0.00, "Yes": 0.05}

# ── Styles ─────────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="2F5496")
PARAM_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_cell(cell, num_format=None):
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center")
    if num_format:
        cell.number_format = num_format


def create_parameters_sheet(wb):
    """Sheet 1: Editable parameters."""
    ws = wb.active
    ws.title = "Parameters"
    ws.sheet_properties.tabColor = "2F5496"

    ws["A1"] = "Monte Carlo Credit Risk Simulation"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")

    ws["A3"] = "Parameter"
    ws["B3"] = "Value"
    ws["C3"] = "Description"
    style_header_row(ws, 3, 3)

    params = [
        ("Number of Simulations", N_SIMS,  "Trials (press F9 to re-simulate)"),
        ("VaR Confidence Level",  0.99,    "For Value-at-Risk calculation"),
    ]
    for i, (name, val, desc) in enumerate(params):
        r = 4 + i
        ws.cell(row=r, column=1, value=name).fill = PARAM_FILL
        c = ws.cell(row=r, column=2, value=val)
        c.fill = INPUT_FILL
        if isinstance(val, float) and val < 1:
            c.number_format = "0.00%"
        else:
            c.number_format = "#,##0"
        ws.cell(row=r, column=3, value=desc)
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = THIN_BORDER

    # Instructions
    ws["A7"] = "How to Use"
    ws["A7"].font = SUBTITLE_FONT
    instructions = [
        "1. Edit per-category correlations in the 'FactorLoadings' sheet",
        "2. Review/edit obligor data in the 'Obligors' sheet",
        "3. Weights (w_region, w_size, w_sector, w_hnwi) are auto-computed via VLOOKUP",
        "4. Press F9 (or Ctrl+Alt+F9) to recalculate = run a new simulation",
        "5. View results in the 'Results' sheet",
        "6. The 'Factors' sheet shows systematic factor draws per trial",
        "7. The 'Simulation' sheet shows the correlated latent variables",
        "8. The 'Defaults' sheet shows binary default indicators (1=default)",
        "9. The 'Losses' sheet shows per-obligor loss amounts",
    ]
    for i, text in enumerate(instructions):
        ws.cell(row=8 + i, column=1, value=text)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 42

    return ws


def create_factor_loadings_sheet(wb):
    """Sheet 2: Per-category factor loadings (editable correlation inputs).

    Layout:
      Region table:  A3:B7   (header + 4 regions)
      Size table:    A10:B12  (header + 2 sizes)
      Sector table:  A15:B19  (header + 4 sectors)
      HNWI table:    A22:B24  (header + 2 values)
    """
    ws = wb.create_sheet("FactorLoadings")
    ws.sheet_properties.tabColor = "7030A0"

    ws["A1"] = "Factor Loadings (Intra-Group Correlations)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")

    # Helper to write a lookup table
    def write_table(start_row, title, data_dict, desc):
        ws.cell(row=start_row, column=1, value=title).font = SUBTITLE_FONT
        ws.cell(row=start_row, column=3, value=desc)
        hr = start_row + 1
        ws.cell(row=hr, column=1, value="Category")
        ws.cell(row=hr, column=2, value="Weight")
        style_header_row(ws, hr, 2)
        for i, (key, val) in enumerate(data_dict.items()):
            r = hr + 1 + i
            ws.cell(row=r, column=1, value=key).fill = PARAM_FILL
            ws.cell(row=r, column=1).border = THIN_BORDER
            c = ws.cell(row=r, column=2, value=val)
            c.number_format = "0.00%"
            c.fill = INPUT_FILL
            c.border = THIN_BORDER
        return hr + 1  # first data row (for VLOOKUP references)

    # Row positions for each table (needed by VLOOKUP in Obligors sheet)
    # Region: rows 3-7 (header=3, data=4-7)
    write_table(3, "Region", REGION_LOADINGS,
                "Correlation contribution from shared region factor")
    # Size: rows 10-12
    write_table(10, "Company Size", SIZE_LOADINGS,
                "Correlation contribution from shared size factor")
    # Sector: rows 15-19
    write_table(15, "Sector", SECTOR_LOADINGS,
                "Correlation contribution from shared sector factor")
    # HNWI: rows 22-24
    write_table(22, "HNWI Status", HNWI_LOADINGS,
                "Correlation contribution from shared HNWI factor")

    # Explanation
    ws["A27"] = "How to Read"
    ws["A27"].font = SUBTITLE_FONT
    explanations = [
        "Each weight represents the share of an obligor's risk variance explained by that factor.",
        "Two obligors sharing a factor contribute that weight to their pairwise correlation.",
        "Example: Two Large US Finance companies → ρ = w_US + w_Large + w_Finance = 0.25+0.12+0.10 = 0.47",
        "Constraint: For each obligor, sum of all weights must be < 1 (remainder = idiosyncratic risk).",
    ]
    for i, text in enumerate(explanations):
        ws.cell(row=28 + i, column=1, value=text)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 55

    return ws


# ── FactorLoadings sheet cell references ──────────────────────────────────
# These constants map to the VLOOKUP ranges in the FactorLoadings sheet.
# Region table: data in A5:B8 (4 rows), Size: A12:B13, Sector: A17:B20, HNWI: A24:B25
FL_REGION_RANGE = "FactorLoadings!$A$5:$B$8"
FL_SIZE_RANGE = "FactorLoadings!$A$12:$B$13"
FL_SECTOR_RANGE = "FactorLoadings!$A$17:$B$20"
FL_HNWI_RANGE = "FactorLoadings!$A$24:$B$25"


def create_obligors_sheet(wb):
    """Sheet 3: Obligor portfolio data with per-obligor factor weights.

    Columns:
      A: ID, B: Name, C: Region, D: Size, E: Sector, F: HNWI,
      G: PD(%), H: EAD(M), I: LGD(%),
      J: w_r (VLOOKUP), K: w_size (VLOOKUP), L: w_s (VLOOKUP), M: w_h (VLOOKUP),
      N: w_idio (=1-J-K-L-M), O: Default Threshold
    """
    ws = wb.create_sheet("Obligors")
    ws.sheet_properties.tabColor = "548235"

    headers = [
        "ID", "Name", "Region", "Size", "Sector", "HNWI",
        "PD (%)", "EAD (M)", "LGD (%)",
        "w_region", "w_size", "w_sector", "w_hnwi", "w_idio",
        "Default Threshold",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for i, ob in enumerate(OBLIGORS):
        r = i + 2
        # ob = (ID, Name, Region, Size, Sector, HNWI, PD%, EAD, LGD%)
        ws.cell(row=r, column=1, value=ob[0])   # ID
        ws.cell(row=r, column=2, value=ob[1])   # Name
        ws.cell(row=r, column=3, value=ob[2])   # Region
        ws.cell(row=r, column=4, value=ob[3])   # Size
        ws.cell(row=r, column=5, value=ob[4])   # Sector
        ws.cell(row=r, column=6, value=ob[5])   # HNWI

        c_pd = ws.cell(row=r, column=7, value=ob[6] / 100)
        c_pd.number_format = "0.00%"
        c_ead = ws.cell(row=r, column=8, value=ob[7])
        c_ead.number_format = "#,##0"
        c_lgd = ws.cell(row=r, column=9, value=ob[8] / 100)
        c_lgd.number_format = "0.00%"

        # Factor weights via VLOOKUP from FactorLoadings sheet
        # J: w_region
        c_wr = ws.cell(row=r, column=10)
        c_wr.value = f"=VLOOKUP(C{r},{FL_REGION_RANGE},2,FALSE)"
        c_wr.number_format = "0.00%"
        # K: w_size
        c_ws = ws.cell(row=r, column=11)
        c_ws.value = f"=VLOOKUP(D{r},{FL_SIZE_RANGE},2,FALSE)"
        c_ws.number_format = "0.00%"
        # L: w_sector
        c_wse = ws.cell(row=r, column=12)
        c_wse.value = f"=VLOOKUP(E{r},{FL_SECTOR_RANGE},2,FALSE)"
        c_wse.number_format = "0.00%"
        # M: w_hnwi
        c_wh = ws.cell(row=r, column=13)
        c_wh.value = f"=VLOOKUP(F{r},{FL_HNWI_RANGE},2,FALSE)"
        c_wh.number_format = "0.00%"
        # N: w_idio = 1 - w_r - w_size - w_s - w_h
        c_wi = ws.cell(row=r, column=14)
        c_wi.value = f"=1-J{r}-K{r}-L{r}-M{r}"
        c_wi.number_format = "0.00%"
        # O: Default threshold = NORM.S.INV(PD)
        c_thr = ws.cell(row=r, column=15)
        c_thr.value = f"=NORM.S.INV(G{r})"
        c_thr.number_format = "0.0000"

        for col in range(1, 16):
            style_data_cell(ws.cell(row=r, column=col))

    # Total exposure
    r_total = len(OBLIGORS) + 3
    ws.cell(row=r_total, column=7, value="Total EAD:").font = Font(bold=True)
    ws.cell(row=r_total, column=8, value=f"=SUM(H2:H{len(OBLIGORS)+1})")
    ws.cell(row=r_total, column=8).number_format = "#,##0"
    ws.cell(row=r_total, column=8).font = Font(bold=True)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 10
    ws.column_dimensions["J"].width = 10
    ws.column_dimensions["K"].width = 10
    ws.column_dimensions["L"].width = 10
    ws.column_dimensions["M"].width = 10
    ws.column_dimensions["N"].width = 10
    ws.column_dimensions["O"].width = 16

    return ws


def create_factors_sheet(wb):
    """Sheet 4: Systematic factor draws (NORM.S.INV(RAND())) for each trial.

    Layout:
      Row 1: headers (Trial 1, Trial 2, ...)
      Col A: Factor labels
      Rows 2-5:  Region factors (US, Europe, Asia ex-Japan, Japan)
      Rows 6-7:  Size factors (Large, SME)
      Rows 8-11: Sector factors (Finance, Manufacturing, Tech, Real Estate)
      Rows 12-13: HNWI factors (No, Yes)
    """
    ws = wb.create_sheet("Factors")
    ws.sheet_properties.tabColor = "BF8F00"

    # Column A labels
    ws.cell(row=1, column=1, value="Factor \\ Trial")
    ws.cell(row=1, column=1).font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL

    factor_labels = []
    for r in REGIONS:
        factor_labels.append(f"Region: {r}")
    for s in SIZES:
        factor_labels.append(f"Size: {s}")
    for s in SECTORS:
        factor_labels.append(f"Sector: {s}")
    for h in HNWI_VALS:
        factor_labels.append(f"HNWI: {h}")

    for i, label in enumerate(factor_labels):
        c = ws.cell(row=i + 2, column=1, value=label)
        c.font = Font(bold=True)
        c.fill = PARAM_FILL
        c.border = THIN_BORDER

    # Trial headers + factor draws
    for j in range(1, N_SIMS + 1):
        col = j + 1
        c = ws.cell(row=1, column=col, value=f"Trial {j}")
        c.font = Font(bold=True, size=9)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")

        # Each factor gets an independent NORM.S.INV(RAND())
        for i in range(len(factor_labels)):
            cell = ws.cell(row=i + 2, column=col)
            cell.value = "=NORM.S.INV(RAND())"
            cell.number_format = "0.0000"
            cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 22
    for j in range(1, N_SIMS + 1):
        ws.column_dimensions[get_column_letter(j + 1)].width = 8

    # Factor row map (for reference in Simulation sheet):
    # Region: rows 2-5  (US=2, Europe=3, Asia ex-Japan=4, Japan=5)
    # Size:   rows 6-7  (Large=6, SME=7)
    # Sector: rows 8-11 (Finance=8, Manufacturing=9, Tech=10, Real Estate=11)
    # HNWI:   rows 12-13 (No=12, Yes=13)

    return ws


def _factor_row_for_obligor(region, size, sector, hnwi):
    """Return the row numbers in the Factors sheet for this obligor's factors."""
    region_row = 2 + REGIONS.index(region)
    size_row = 2 + len(REGIONS) + SIZES.index(size)
    sector_row = 2 + len(REGIONS) + len(SIZES) + SECTORS.index(sector)
    hnwi_row = 2 + len(REGIONS) + len(SIZES) + len(SECTORS) + HNWI_VALS.index(hnwi)
    return region_row, size_row, sector_row, hnwi_row


def create_simulation_sheet(wb):
    """Sheet 5: Correlated latent variable Z for each obligor x trial.

    Z_i = sqrt(w_r)*F_region + sqrt(w_size)*F_size + sqrt(w_s)*F_sector
          + sqrt(w_h)*F_hnwi + sqrt(w_idio)*NORM.S.INV(RAND())

    Per-obligor weights come from Obligors sheet columns J-N.
    """
    ws = wb.create_sheet("Simulation")
    ws.sheet_properties.tabColor = "C55A11"

    # Column A: obligor IDs
    ws.cell(row=1, column=1, value="Obligor \\ Trial")
    ws.cell(row=1, column=1).font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL

    for i, ob in enumerate(OBLIGORS):
        c = ws.cell(row=i + 2, column=1, value=ob[0])
        c.font = Font(bold=True)
        c.fill = PARAM_FILL
        c.border = THIN_BORDER

    # Trial headers
    for j in range(1, N_SIMS + 1):
        col = j + 1
        c = ws.cell(row=1, column=col, value=f"Trial {j}")
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

    # Formulas - now using per-obligor weights from Obligors sheet
    for i, ob in enumerate(OBLIGORS):
        r_row = i + 2
        # ob = (ID, Name, Region, Size, Sector, HNWI, ...)
        region_row, size_row, sector_row, hnwi_row = _factor_row_for_obligor(
            ob[2], ob[3], ob[4], ob[5]
        )

        # Obligors sheet weight columns (absolute row reference per obligor):
        # J=w_region, K=w_size, L=w_sector, M=w_hnwi, N=w_idio
        wr = f"Obligors!$J${r_row}"
        wz = f"Obligors!$K${r_row}"
        ws_ref = f"Obligors!$L${r_row}"
        wh = f"Obligors!$M${r_row}"
        wi = f"Obligors!$N${r_row}"

        for j in range(1, N_SIMS + 1):
            col = j + 1
            col_letter = get_column_letter(col)

            f_region = f"Factors!{col_letter}{region_row}"
            f_size = f"Factors!{col_letter}{size_row}"
            f_sector = f"Factors!{col_letter}{sector_row}"
            f_hnwi = f"Factors!{col_letter}{hnwi_row}"

            formula = (
                f"=SQRT({wr})*{f_region}"
                f"+SQRT({wz})*{f_size}"
                f"+SQRT({ws_ref})*{f_sector}"
                f"+SQRT({wh})*{f_hnwi}"
                f"+SQRT({wi})*NORM.S.INV(RAND())"
            )
            cell = ws.cell(row=r_row, column=col, value=formula)
            cell.number_format = "0.0000"

    ws.column_dimensions["A"].width = 12
    return ws


def create_defaults_sheet(wb):
    """Sheet 5: Binary default indicators.
    =IF(Simulation!cell < Obligors!$I$row, 1, 0)
    """
    ws = wb.create_sheet("Defaults")
    ws.sheet_properties.tabColor = "FF0000"

    ws.cell(row=1, column=1, value="Obligor \\ Trial")
    ws.cell(row=1, column=1).font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL

    for i, ob in enumerate(OBLIGORS):
        c = ws.cell(row=i + 2, column=1, value=ob[0])
        c.font = Font(bold=True)
        c.fill = PARAM_FILL
        c.border = THIN_BORDER

    for j in range(1, N_SIMS + 1):
        col = j + 1
        col_letter = get_column_letter(col)
        c = ws.cell(row=1, column=col, value=f"Trial {j}")
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

        for i in range(len(OBLIGORS)):
            r = i + 2
            # Default if Z < threshold (column O in Obligors)
            formula = f"=IF(Simulation!{col_letter}{r}<Obligors!$O${r},1,0)"
            cell = ws.cell(row=r, column=col, value=formula)
            cell.number_format = "0"

    ws.column_dimensions["A"].width = 12
    return ws


def create_losses_sheet(wb):
    """Sheet 6: Loss per obligor per trial = Default * EAD * LGD."""
    ws = wb.create_sheet("Losses")
    ws.sheet_properties.tabColor = "7030A0"

    ws.cell(row=1, column=1, value="Obligor \\ Trial")
    ws.cell(row=1, column=1).font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL

    for i, ob in enumerate(OBLIGORS):
        c = ws.cell(row=i + 2, column=1, value=ob[0])
        c.font = Font(bold=True)
        c.fill = PARAM_FILL
        c.border = THIN_BORDER

    # Portfolio loss row
    loss_total_row = len(OBLIGORS) + 3
    ws.cell(row=loss_total_row, column=1, value="Portfolio Loss")
    ws.cell(row=loss_total_row, column=1).font = Font(bold=True, color="FF0000")

    for j in range(1, N_SIMS + 1):
        col = j + 1
        col_letter = get_column_letter(col)
        c = ws.cell(row=1, column=col, value=f"Trial {j}")
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

        for i in range(len(OBLIGORS)):
            r = i + 2
            # Loss = Default_indicator * EAD * LGD (columns H and I in Obligors)
            formula = f"=Defaults!{col_letter}{r}*Obligors!$H${r}*Obligors!$I${r}"
            cell = ws.cell(row=r, column=col, value=formula)
            cell.number_format = "#,##0.0"

        # Sum column for portfolio loss
        first_row = 2
        last_row = len(OBLIGORS) + 1
        formula = f"=SUM({col_letter}{first_row}:{col_letter}{last_row})"
        cell = ws.cell(row=loss_total_row, column=col, value=formula)
        cell.number_format = "#,##0.0"
        cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 16
    return ws


def create_results_sheet(wb):
    """Sheet 7: Summary statistics and histogram."""
    ws = wb.create_sheet("Results")
    ws.sheet_properties.tabColor = "00B050"

    loss_total_row = len(OBLIGORS) + 3  # row in Losses sheet with portfolio totals
    last_col_letter = get_column_letter(N_SIMS + 1)

    # ── Title ──
    ws["A1"] = "Monte Carlo Simulation Results"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    # ── Summary Statistics ──
    ws["A3"] = "Summary Statistics"
    ws["A3"].font = SUBTITLE_FONT

    stats = [
        ("Expected Loss (Mean)", f"=AVERAGE(Losses!B{loss_total_row}:{last_col_letter}{loss_total_row})", "#,##0.0"),
        ("Loss Std Dev", f"=STDEV(Losses!B{loss_total_row}:{last_col_letter}{loss_total_row})", "#,##0.0"),
        ("VaR (99%)", f"=PERCENTILE(Losses!B{loss_total_row}:{last_col_letter}{loss_total_row},Parameters!B5)", "#,##0.0"),
        ("Expected Shortfall (CVaR)",
         f"=AVERAGEIF(Losses!B{loss_total_row}:{last_col_letter}{loss_total_row},\">=\"&B7)",
         "#,##0.0"),
        ("Max Loss", f"=MAX(Losses!B{loss_total_row}:{last_col_letter}{loss_total_row})", "#,##0.0"),
        ("Min Loss", f"=MIN(Losses!B{loss_total_row}:{last_col_letter}{loss_total_row})", "#,##0.0"),
        ("Median Loss", f"=MEDIAN(Losses!B{loss_total_row}:{last_col_letter}{loss_total_row})", "#,##0.0"),
        ("Total Portfolio EAD", f"=Obligors!H{len(OBLIGORS)+3}", "#,##0"),
        ("Expected Loss Rate", f"=B5/B12", "0.00%"),
    ]

    ws["A4"] = "Metric"
    ws["B4"] = "Value"
    ws["C4"] = "Unit"
    style_header_row(ws, 4, 3)

    units = ["M", "M", "M", "M", "M", "M", "M", "M", "%"]
    for i, (label, formula, fmt) in enumerate(stats):
        r = 5 + i
        ws.cell(row=r, column=1, value=label).fill = PARAM_FILL
        c = ws.cell(row=r, column=2, value=formula)
        c.number_format = fmt
        c.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        ws.cell(row=r, column=3, value=units[i])
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = THIN_BORDER

    # ── Default Count Statistics ──
    ws["A16"] = "Default Count Statistics"
    ws["A16"].font = SUBTITLE_FONT

    ws["A17"] = "Metric"
    ws["B17"] = "Value"
    style_header_row(ws, 17, 2)

    # Count defaults per trial (sum of Defaults sheet column)
    # We need a helper row. Use row 18-22 for stats based on default counts.
    # Average number of defaults per trial
    first_col = "B"
    # We'll compute from Defaults sheet: each trial column sum
    # Use a trick: SUMPRODUCT across the Defaults range / N_SIMS for avg defaults
    n_ob = len(OBLIGORS)
    ws.cell(row=18, column=1, value="Avg Defaults per Trial").fill = PARAM_FILL
    ws.cell(row=18, column=2,
            value=f"=SUMPRODUCT(Defaults!B2:{last_col_letter}{n_ob+1})/{N_SIMS}")
    ws.cell(row=18, column=2).number_format = "0.00"

    ws.cell(row=19, column=1, value="Avg Default Rate").fill = PARAM_FILL
    ws.cell(row=19, column=2, value=f"=B18/{n_ob}")
    ws.cell(row=19, column=2).number_format = "0.00%"

    for r in range(18, 20):
        for col in range(1, 3):
            ws.cell(row=r, column=col).border = THIN_BORDER

    # ── Loss Distribution (Histogram Data) ──
    ws["A22"] = "Loss Distribution"
    ws["A22"].font = SUBTITLE_FONT

    ws["A23"] = "Bucket Upper"
    ws["B23"] = "Frequency"
    ws["C23"] = "Cumulative %"
    style_header_row(ws, 23, 3)

    # Bucket boundaries: from 0 to Max Loss in N_BUCKETS steps
    # Bucket width = MAX(losses) / N_BUCKETS, but we use a fixed reasonable range
    # Use bin edges based on a formula referencing the max loss
    for b in range(N_BUCKETS):
        r = 24 + b
        # Bucket upper bound = (b+1) * MaxLoss / N_BUCKETS
        ws.cell(row=r, column=1,
                value=f"=({b + 1})*B9/{N_BUCKETS}")
        ws.cell(row=r, column=1).number_format = "#,##0.0"
        ws.cell(row=r, column=1).border = THIN_BORDER

    # FREQUENCY array - in openpyxl we write as CSE (Ctrl+Shift+Enter) array formula
    # FREQUENCY(data_array, bins_array)
    freq_start = 24
    freq_end = 24 + N_BUCKETS - 1
    data_range = f"Losses!B{loss_total_row}:{last_col_letter}{loss_total_row}"
    bins_range = f"A{freq_start}:A{freq_end}"

    for b in range(N_BUCKETS):
        r = 24 + b
        # Use COUNTIFS instead of FREQUENCY for compatibility (no CSE needed)
        if b == 0:
            formula = f'=COUNTIF({data_range},"<="&A{r})'
        else:
            formula = f'=COUNTIFS({data_range},"<="&A{r})-COUNTIFS({data_range},"<="&A{r - 1})'
        ws.cell(row=r, column=2, value=formula)
        ws.cell(row=r, column=2).number_format = "#,##0"
        ws.cell(row=r, column=2).border = THIN_BORDER

        # Cumulative %
        if b == 0:
            ws.cell(row=r, column=3, value=f"=B{r}/{N_SIMS}")
        else:
            ws.cell(row=r, column=3, value=f"=C{r - 1}+B{r}/{N_SIMS}")
        ws.cell(row=r, column=3).number_format = "0.0%"
        ws.cell(row=r, column=3).border = THIN_BORDER

    # ── Histogram Chart ──
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Portfolio Loss Distribution"
    chart.x_axis.title = "Loss Amount (M)"
    chart.y_axis.title = "Frequency"
    chart.width = 24
    chart.height = 14

    cats = Reference(ws, min_col=1, min_row=freq_start, max_row=freq_end)
    data = Reference(ws, min_col=2, min_row=23, max_row=freq_end)  # include header
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4

    # Style the bars
    series = chart.series[0]
    series.graphicalProperties.solidFill = "2F5496"
    chart.legend = None

    ws.add_chart(chart, "E4")

    # ── Cumulative Distribution Chart ──
    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 10
    chart2.title = "Cumulative Loss Distribution"
    chart2.x_axis.title = "Loss Amount (M)"
    chart2.y_axis.title = "Cumulative Probability"
    chart2.width = 24
    chart2.height = 14

    from openpyxl.chart import LineChart
    chart2 = LineChart()
    chart2.title = "Cumulative Loss Distribution"
    chart2.x_axis.title = "Loss Amount (M)"
    chart2.y_axis.title = "Cumulative Probability"
    chart2.width = 24
    chart2.height = 14

    data2 = Reference(ws, min_col=3, min_row=23, max_row=freq_end)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats)
    series2 = chart2.series[0]
    series2.graphicalProperties.line.solidFill = "C55A11"
    chart2.legend = None

    ws.add_chart(chart2, "E22")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16

    return ws


def main():
    wb = openpyxl.Workbook()

    print("Creating Parameters sheet...")
    create_parameters_sheet(wb)

    print("Creating FactorLoadings sheet...")
    create_factor_loadings_sheet(wb)

    print("Creating Obligors sheet...")
    create_obligors_sheet(wb)

    print("Creating Factors sheet...")
    create_factors_sheet(wb)

    print("Creating Simulation sheet...")
    create_simulation_sheet(wb)

    print("Creating Defaults sheet...")
    create_defaults_sheet(wb)

    print("Creating Losses sheet...")
    create_losses_sheet(wb)

    print("Creating Results sheet...")
    create_results_sheet(wb)

    output_path = "monte_carlo_credit_risk.xlsx"
    wb.save(output_path)
    print(f"\nDone! Saved to: {output_path}")
    print(f"  - {len(OBLIGORS)} obligors x {N_SIMS} trials")
    print("  - Open in Excel and press F9 to recalculate (new simulation)")
    print("  - Adjust per-category correlations in the 'FactorLoadings' sheet")


if __name__ == "__main__":
    main()
