#!/usr/bin/env python3
"""
HHI vs Loss Rate - Excel with Pre-computed Simulation + Formulas

Generates an Excel file where:
  - Simulation data is pre-computed by Python (numpy) and written as values
  - Results sheet uses Excel formulas (AVERAGE, STDEV, PERCENTILE) on the data
  - Weights sheet uses Excel formulas (SUMPRODUCT for HHI, SUM for totals)
  - Charts auto-render from formula results

The user can:
  1. Open in Excel and immediately see results (pre-computed data)
  2. Edit weights, then F9 to see Results formulas update
  3. Understand every formula used

Conditions: correlation=0, PD=AA (0.03%), LGD=100%, 10 obligors
"""

import numpy as np
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration ────────────────────────────────────────────────────────
N_OBLIGORS = 10
N_SIMS = 500          # trials (keeps Excel responsive)
N_SCENARIOS = 6       # different weight distributions to compare
DEFAULT_PD = 0.0003   # 0.03% (AA)
DEFAULT_LGD = 1.0     # 100%

# ── Styles ───────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="2F5496")
PARAM_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
WARN_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_cell(cell, num_format=None):
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center")
    if num_format:
        cell.number_format = num_format


# ── Predefined weight scenarios ──────────────────────────────────────────
SCENARIOS = [
    ("均等配分",        [10, 10, 10, 10, 10, 10, 10, 10, 10, 10]),
    ("やや集中",       [20, 15, 12, 11, 10,  8,  8,  7,  5,  4]),
    ("中程度集中",     [30, 20, 15, 10,  8,  5,  5,  3,  2,  2]),
    ("高集中",         [50, 15, 10,  7,  5,  4,  3,  3,  2,  1]),
    ("極端な集中",     [70, 10,  5,  4,  3,  2,  2,  2,  1,  1]),
    ("1社集中",        [91,  1,  1,  1,  1,  1,  1,  1,  1,  1]),
]


def run_monte_carlo(weights, pd=DEFAULT_PD, lgd=DEFAULT_LGD, n_sims=N_SIMS):
    """Run Monte Carlo in numpy. Returns (defaults_matrix, loss_array)."""
    w = np.array(weights) / 100.0
    defaults = (np.random.random((n_sims, len(w))) < pd).astype(int)
    losses = (defaults * w) @ np.ones(len(w)) * lgd  # = sum(D_i * w_i) * LGD
    # More precisely: losses = defaults @ w * lgd
    losses = defaults @ w * lgd
    return defaults, losses


def create_parameters_sheet(wb):
    ws = wb.active
    ws.title = "Parameters"
    ws.sheet_properties.tabColor = "2F5496"

    ws["A1"] = "HHI vs 損失率 Monte Carlo分析"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    ws["A3"] = "パラメータ"
    ws["B3"] = "値"
    ws["C3"] = "説明"
    style_header_row(ws, 3, 3)

    params = [
        ("PD (デフォルト確率)", DEFAULT_PD, "0.00%", "AA格の年間PD"),
        ("LGD (損失率)", DEFAULT_LGD, "0%", "デフォルト時の損失割合"),
        ("シミュレーション回数", N_SIMS, "#,##0", f"試行回数 ({N_SIMS}回)"),
        ("債務者数", N_OBLIGORS, "#,##0", "ポートフォリオの企業数"),
    ]
    for i, (name, val, fmt, desc) in enumerate(params):
        r = 4 + i
        ws.cell(row=r, column=1, value=name).fill = PARAM_FILL
        c = ws.cell(row=r, column=2, value=val)
        c.number_format = fmt
        c.fill = INPUT_FILL
        ws.cell(row=r, column=3, value=desc)
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = THIN_BORDER

    ws["A9"] = "構造"
    ws["A9"].font = SUBTITLE_FONT
    notes = [
        "Weightsシート: 各シナリオのウェイト + HHI計算式 (=SUMPRODUCT)",
        "Sim_N シート: Pythonで実行したMC結果（0/1のデフォルトデータ）",
        "  → 最下行: ポートフォリオ損失 = Excel数式 =SUMPRODUCT(defaults, weights)*LGD",
        "Resultsシート: Excel数式で集計（=AVERAGE, =STDEV, =PERCENTILE, =MAX）",
        "  → チャート3つが数式結果から自動生成",
    ]
    for i, text in enumerate(notes):
        ws.cell(row=10 + i, column=1, value=text)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 36


def create_weights_sheet(wb):
    ws = wb.create_sheet("Weights")
    ws.sheet_properties.tabColor = "7030A0"

    ws["A1"] = "債務者ウェイト配分（各シナリオ）"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:M1")

    # Headers
    headers = (["シナリオ"]
               + [f"債務者{i+1}" for i in range(N_OBLIGORS)]
               + ["合計", "HHI", "HHI数式"])
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, len(headers))

    first_w = get_column_letter(2)
    last_w = get_column_letter(N_OBLIGORS + 1)

    for s_idx, (name, pcts) in enumerate(SCENARIOS):
        r = 4 + s_idx
        ws.cell(row=r, column=1, value=name).fill = PARAM_FILL
        ws.cell(row=r, column=1).border = THIN_BORDER

        for j, pct in enumerate(pcts):
            c = ws.cell(row=r, column=2 + j, value=pct / 100)
            c.number_format = "0.0%"
            c.fill = INPUT_FILL
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center")

        # SUM formula
        sum_col = N_OBLIGORS + 2
        c = ws.cell(row=r, column=sum_col)
        c.value = f"=SUM({first_w}{r}:{last_w}{r})"
        c.number_format = "0.0%"
        style_cell(c, "0.0%")

        # HHI formula
        hhi_col = N_OBLIGORS + 3
        c = ws.cell(row=r, column=hhi_col)
        c.value = f"=SUMPRODUCT({first_w}{r}:{last_w}{r},{first_w}{r}:{last_w}{r})"
        c.number_format = "0.0000"
        style_cell(c, "0.0000")

        # Show formula as text for reference
        formula_col = N_OBLIGORS + 4
        ws.cell(row=r, column=formula_col,
                value=f'=SUMPRODUCT(B{r}:K{r},B{r}:K{r})')

    ws.column_dimensions["A"].width = 16
    for c in range(2, N_OBLIGORS + 5):
        ws.column_dimensions[get_column_letter(c)].width = 11


def create_simulation_sheet(wb, scenario_idx, defaults_matrix, losses):
    """Write pre-computed simulation data + Excel formulas for portfolio loss."""
    name = SCENARIOS[scenario_idx][0]
    ws = wb.create_sheet(f"Sim_{scenario_idx+1}_{name}")
    ws.sheet_properties.tabColor = "C55A11"

    s_row = 4 + scenario_idx  # weight row in Weights sheet

    # Column A: labels
    ws.cell(row=1, column=1, value="債務者 \\ Trial")
    ws.cell(row=1, column=1).font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL

    for i in range(N_OBLIGORS):
        c = ws.cell(row=i + 2, column=1, value=f"債務者{i+1}")
        c.font = Font(bold=True)
        c.fill = PARAM_FILL
        c.border = THIN_BORDER

    # Formula description row
    desc_row = N_OBLIGORS + 2
    ws.cell(row=desc_row, column=1, value="数式")
    ws.cell(row=desc_row, column=1).font = Font(bold=True, italic=True, color="808080")

    # Portfolio loss row
    loss_row = N_OBLIGORS + 3
    ws.cell(row=loss_row, column=1, value="Portfolio Loss")
    ws.cell(row=loss_row, column=1).font = Font(bold=True, color="FF0000")
    ws.cell(row=loss_row, column=1).border = THIN_BORDER

    # Loss formula description row
    loss_desc_row = N_OBLIGORS + 4
    ws.cell(row=loss_desc_row, column=1, value="損失の数式")
    ws.cell(row=loss_desc_row, column=1).font = Font(bold=True, italic=True, color="808080")

    first_r = 2
    last_r = N_OBLIGORS + 1
    w_first = get_column_letter(2)
    w_last = get_column_letter(N_OBLIGORS + 1)

    for j in range(N_SIMS):
        col = j + 2
        col_letter = get_column_letter(col)

        # Trial header
        c = ws.cell(row=1, column=col, value=f"Trial {j+1}")
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")

        # Pre-computed default indicators (0 or 1)
        for i in range(N_OBLIGORS):
            cell = ws.cell(row=i + 2, column=col, value=int(defaults_matrix[j, i]))
            cell.number_format = "0"

        # Formula description (only first trial)
        if j == 0:
            ws.cell(row=desc_row, column=col,
                    value="↑ Pythonで計算済み: IF(RAND()<PD, 1, 0)")
            ws.cell(row=desc_row, column=col).font = Font(italic=True, color="808080")

        # Portfolio loss: Excel formula =SUMPRODUCT(defaults, weights)*LGD
        formula = (
            f"=SUMPRODUCT({col_letter}{first_r}:{col_letter}{last_r},"
            f"Weights!{w_first}${s_row}:{w_last}${s_row})"
            f"*Parameters!$B$5"
        )
        cell = ws.cell(row=loss_row, column=col, value=formula)
        cell.number_format = "0.00%"
        cell.font = Font(bold=True)

        # Formula text (only first trial)
        if j == 0:
            ws.cell(row=loss_desc_row, column=col,
                    value=f"=SUMPRODUCT({col_letter}2:{col_letter}{last_r},"
                          f"Weights!B$4:K$4)*Parameters!$B$5")
            ws.cell(row=loss_desc_row, column=col).font = Font(italic=True, color="808080")

    ws.column_dimensions["A"].width = 16
    return ws


def create_results_sheet(wb):
    """Results with ALL Excel formulas + formula text reference."""
    ws = wb.create_sheet("Results")
    ws.sheet_properties.tabColor = "00B050"

    ws["A1"] = "HHI vs 損失率 シミュレーション結果"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")

    ws["A2"] = "※ 全ての値はExcel数式で計算されています（右の列に数式テキスト表示）"
    ws["A2"].font = Font(italic=True, color="FF0000")

    loss_row = N_OBLIGORS + 3
    last_col = get_column_letter(N_SIMS + 1)

    # ── Summary Table ────────────────────────────────────────────────────
    ws["A4"] = "シナリオ別 損失指標"
    ws["A4"].font = SUBTITLE_FONT

    headers = [
        "シナリオ", "HHI\n(数式)",
        "期待損失率\n=AVERAGE()",
        "標準偏差\n=STDEV()",
        "理論値 σ\n=SQRT()",
        "VaR 99%\n=PERCENTILE(,0.99)",
        "VaR 99.9%\n=PERCENTILE(,0.999)",
        "最大損失率\n=MAX()",
        "最大ウェイト\n=MAX()",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=5, column=c, value=h)
    style_header_row(ws, 5, len(headers))

    # Formula text column headers (columns K onwards)
    formula_headers = [
        "HHI数式", "E[L]数式", "σ数式", "理論σ数式",
        "VaR99%数式", "VaR99.9%数式", "MAX数式",
    ]
    for c, h in enumerate(formula_headers):
        cell = ws.cell(row=5, column=11 + c, value=h)
        cell.font = Font(bold=True, color="808080", size=9)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    hhi_col_letter = get_column_letter(N_OBLIGORS + 3)
    w_first = get_column_letter(2)
    w_last = get_column_letter(N_OBLIGORS + 1)

    for s_idx, (name, _) in enumerate(SCENARIOS):
        r = 6 + s_idx
        sim_sheet = f"Sim_{s_idx+1}_{name}"
        s_weight_row = 4 + s_idx

        loss_range = f"'{sim_sheet}'!B{loss_row}:{last_col}{loss_row}"

        # A: Scenario name
        ws.cell(row=r, column=1, value=name).fill = PARAM_FILL
        style_cell(ws.cell(row=r, column=1))

        # B: HHI (formula)
        f_hhi = f"=Weights!{hhi_col_letter}{s_weight_row}"
        c = ws.cell(row=r, column=2, value=f_hhi)
        style_cell(c, "0.0000")

        # C: Expected Loss = AVERAGE
        f_el = f"=AVERAGE({loss_range})"
        c = ws.cell(row=r, column=3, value=f_el)
        style_cell(c, "0.0000%")

        # D: Std Dev = STDEV
        f_sd = f"=STDEV({loss_range})"
        c = ws.cell(row=r, column=4, value=f_sd)
        style_cell(c, "0.0000%")

        # E: Theoretical σ
        f_theo = f"=SQRT(B{r}*Parameters!$B$4*(1-Parameters!$B$4))*Parameters!$B$5"
        c = ws.cell(row=r, column=5, value=f_theo)
        style_cell(c, "0.0000%")

        # F: VaR 99%
        f_v99 = f"=PERCENTILE({loss_range},0.99)"
        c = ws.cell(row=r, column=6, value=f_v99)
        style_cell(c, "0.00%")

        # G: VaR 99.9%
        f_v999 = f"=PERCENTILE({loss_range},0.999)"
        c = ws.cell(row=r, column=7, value=f_v999)
        style_cell(c, "0.00%")

        # H: Max Loss
        f_max = f"=MAX({loss_range})"
        c = ws.cell(row=r, column=8, value=f_max)
        style_cell(c, "0.00%")

        # I: Max Weight
        f_mw = f"=MAX(Weights!{w_first}{s_weight_row}:{w_last}{s_weight_row})"
        c = ws.cell(row=r, column=9, value=f_mw)
        style_cell(c, "0.0%")

        # K-Q: Formula text (visible as reference)
        formulas_text = [f_hhi, f_el, f_sd, f_theo, f_v99, f_v999, f_max]
        for fi, ft in enumerate(formulas_text):
            cell = ws.cell(row=r, column=11 + fi, value=ft)
            cell.font = Font(color="808080", size=9)

    last_data_row = 5 + N_SCENARIOS

    # ── Theoretical E[L] ─────────────────────────────────────────────────
    ws.cell(row=last_data_row + 2, column=1, value="理論値 E[L]").font = SUBTITLE_FONT
    c = ws.cell(row=last_data_row + 2, column=2,
                value="=Parameters!$B$4*Parameters!$B$5")
    style_cell(c, "0.0000%")
    ws.cell(row=last_data_row + 2, column=3,
            value="= PD × LGD（HHIに無関係 → 期待損失は集中度に依存しない）")

    # ── Formula Summary ──────────────────────────────────────────────────
    fs_row = last_data_row + 4
    ws.cell(row=fs_row, column=1, value="数式の一覧").font = SUBTITLE_FONT

    formula_table = [
        ("計算項目", "Excel数式", "意味"),
        ("HHI", "=SUMPRODUCT(w, w)", "Σ(w_i²) — ウェイトの二乗和 = 集中度"),
        ("デフォルト判定", "IF(RAND()<PD, 1, 0)", "各債務者が独立にPDの確率でデフォルト"),
        ("ポートフォリオ損失", "=SUMPRODUCT(D, w)*LGD", "Σ(D_i × w_i) × LGD"),
        ("期待損失率 E[L]", "=AVERAGE(損失)", "= PD × LGD（常に0.03%）"),
        ("標準偏差 σ", "=STDEV(損失)", "≈ √(HHI × PD × (1-PD))"),
        ("理論値 σ", "=SQRT(HHI*PD*(1-PD))*LGD", "解析解（シミュレーションと一致するはず）"),
        ("VaR 99.9%", "=PERCENTILE(損失, 0.999)", "上位0.1%の損失境界"),
        ("最大損失率", "=MAX(損失)", "全試行中の最悪ケース"),
    ]
    for i, row_data in enumerate(formula_table):
        r = fs_row + 1 + i
        for j, val in enumerate(row_data):
            cell = ws.cell(row=r, column=1 + j, value=val)
            cell.border = THIN_BORDER
            if i == 0:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            else:
                cell.fill = PARAM_FILL if j == 0 else PatternFill()

    # ── Chart 1: σ vs theoretical ────────────────────────────────────────
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "標準偏差: シミュレーション =STDEV() vs 理論値 =SQRT()"
    chart1.x_axis.title = "シナリオ（集中度↑）"
    chart1.y_axis.title = "標準偏差"
    chart1.y_axis.numFmt = "0.00%"
    chart1.width = 22
    chart1.height = 13

    cats = Reference(ws, min_col=1, min_row=6, max_row=last_data_row)
    chart1.add_data(Reference(ws, min_col=4, min_row=5, max_row=last_data_row),
                    titles_from_data=True)
    chart1.add_data(Reference(ws, min_col=5, min_row=5, max_row=last_data_row),
                    titles_from_data=True)
    chart1.set_categories(cats)
    chart1.series[0].graphicalProperties.solidFill = "FF9800"
    chart1.series[1].graphicalProperties.solidFill = "FF0000"

    ws.add_chart(chart1, "A" + str(fs_row + 12))

    # ── Chart 2: VaR pitfall ─────────────────────────────────────────────
    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 10
    chart2.title = "VaRの落とし穴: =PERCENTILE(,0.999) vs =MAX()"
    chart2.x_axis.title = "シナリオ（集中度↑）"
    chart2.y_axis.title = "損失率"
    chart2.y_axis.numFmt = "0%"
    chart2.width = 22
    chart2.height = 13

    chart2.add_data(Reference(ws, min_col=7, min_row=5, max_row=last_data_row),
                    titles_from_data=True)
    chart2.add_data(Reference(ws, min_col=8, min_row=5, max_row=last_data_row),
                    titles_from_data=True)
    chart2.set_categories(cats)
    chart2.series[0].graphicalProperties.solidFill = "E91E63"
    chart2.series[1].graphicalProperties.solidFill = "9C27B0"

    ws.add_chart(chart2, "A" + str(fs_row + 28))

    # ── Chart 3: All metrics ─────────────────────────────────────────────
    chart3 = BarChart()
    chart3.type = "col"
    chart3.style = 10
    chart3.title = "全指標比較: =AVERAGE() / =STDEV() / =PERCENTILE() / =MAX()"
    chart3.x_axis.title = "シナリオ（集中度↑）"
    chart3.y_axis.title = "損失率"
    chart3.y_axis.numFmt = "0%"
    chart3.width = 22
    chart3.height = 13

    for ci in [3, 4, 7, 8]:
        chart3.add_data(Reference(ws, min_col=ci, min_row=5, max_row=last_data_row),
                        titles_from_data=True)
    chart3.set_categories(cats)
    for i, color in enumerate(["2196F3", "FF9800", "E91E63", "9C27B0"]):
        chart3.series[i].graphicalProperties.solidFill = color

    ws.add_chart(chart3, "A" + str(fs_row + 44))

    ws.column_dimensions["A"].width = 20
    for c in range(2, 10):
        ws.column_dimensions[get_column_letter(c)].width = 16
    for c in range(11, 18):
        ws.column_dimensions[get_column_letter(c)].width = 30


def main():
    np.random.seed(42)
    wb = openpyxl.Workbook()

    print("Creating Parameters sheet...")
    create_parameters_sheet(wb)

    print("Creating Weights sheet...")
    create_weights_sheet(wb)

    print(f"Running Monte Carlo ({N_SCENARIOS} scenarios x {N_SIMS} trials)...")
    for s_idx, (name, pcts) in enumerate(SCENARIOS):
        weights = np.array(pcts, dtype=float)
        hhi = np.sum((weights / 100) ** 2)

        defaults, losses = run_monte_carlo(weights)
        print(f"  Scenario {s_idx+1}: {name:8s}  HHI={hhi:.4f}  "
              f"E[L]={np.mean(losses)*100:.4f}%  "
              f"σ={np.std(losses)*100:.4f}%  "
              f"max={np.max(losses)*100:.2f}%")

        create_simulation_sheet(wb, s_idx, defaults, losses)

    print("Creating Results sheet (Excel formulas)...")
    create_results_sheet(wb)

    output_path = "hhi_loss_formula.xlsx"
    wb.save(output_path)
    print(f"\nDone! Saved to: {output_path}")
    print(f"\nExcel数式の構造:")
    print(f"  Weights!M列    : =SUMPRODUCT(w,w)        → HHI計算")
    print(f"  Sim各シート最下行: =SUMPRODUCT(D,w)*LGD     → 損失計算")
    print(f"  Results!C列    : =AVERAGE(損失)           → 期待損失")
    print(f"  Results!D列    : =STDEV(損失)             → 標準偏差")
    print(f"  Results!E列    : =SQRT(HHI*PD*(1-PD))*LGD → 理論値σ")
    print(f"  Results!F列    : =PERCENTILE(損失,0.99)    → VaR 99%")
    print(f"  Results!G列    : =PERCENTILE(損失,0.999)   → VaR 99.9%")
    print(f"  Results!H列    : =MAX(損失)               → 最大損失")


if __name__ == "__main__":
    main()
